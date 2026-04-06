"""Export utilities for CSV, XLSX, and PDF reports."""

import csv
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_to_csv(headers: list[str], rows: list[dict], key_map: list[str] | None = None) -> str:
    """Export data to CSV string."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    keys = key_map or headers
    for row in rows:
        writer.writerow([row.get(k, "") for k in keys])
    return output.getvalue()


def export_to_xlsx(
    headers: list[str],
    rows: list[dict],
    key_map: list[str] | None = None,
    sheet_name: str = "Reporte",
    title: str | None = None,
) -> bytes:
    """Export data to XLSX bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="EA7704", end_color="EA7704", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)

    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = title_font
        start_row = 3

    keys = key_map or headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, key in enumerate(keys, 1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Right-align numbers
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00"

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
